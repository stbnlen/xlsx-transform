import io

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REPORT_SHEET = "REPORTES_GESTIONES_TODAS_ETAPAS"
PIVOT_SHEET = "Hoja1"
DATA_SHEET = "DATOS"
NEW_RECORDS_SHEET = "REGISTROS_NUEVOS"


def _find_cae_column(df: pd.DataFrame, name: str):
    """Find a column by case-insensitive, whitespace-trimmed name."""
    target = str(name).strip().lower()
    for col in df.columns:
        if str(col).strip().lower() == target:
            return col
    return None


def resolve_cae_pivot_columns(df: pd.DataFrame) -> dict:
    """Resolve the columns required for the CAE pivot table.

    Raises KeyError listing any required column that is missing.
    """
    resolved = {
        "Mes": _find_cae_column(df, "Mes"),
        "ETAPA SATCHMO": _find_cae_column(df, "ETAPA SATCHMO"),
        "PRODUCTO": _find_cae_column(df, "PRODUCTO"),
        "FECHA CREA": _find_cae_column(df, "FECHA CREA"),
    }
    missing = [name for name, col in resolved.items() if col is None]
    if missing:
        raise KeyError("Columnas faltantes: " + ", ".join(missing))
    return resolved


def _add_formatted_table(ws, df: pd.DataFrame, display_name: str) -> None:
    """Add a styled Excel table over the df range and autofit column widths."""
    end_col = get_column_letter(len(df.columns))
    end_row = len(df) + 1
    table = Table(displayName=display_name, ref=f"A1:{end_col}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium7",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    for idx, column in enumerate(df.columns, start=1):
        width = max(
            len(str(column)),
            int(df[column].astype(str).str.len().max()),
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, 255)


def write_cae_nuevos_excel(df_nuevos: pd.DataFrame) -> bytes:
    """Write new records to an xlsx (REPORT_SHEET) and return the bytes."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_nuevos.to_excel(writer, sheet_name=REPORT_SHEET, index=False)
        _add_formatted_table(writer.sheets[REPORT_SHEET], df_nuevos, "RegistrosNuevos")
    return output.getvalue()


def append_new_records_to_report(report_bytes: bytes, df_nuevos: pd.DataFrame) -> bytes:
    """Append new records to the main data sheet (REPORTES_GESTIONES_TODAS_ETAPAS)
    so the existing pivot table (Hoja3) reflects them when refreshed.
    """
    wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
    ws = wb[REPORT_SHEET]

    # Find the table in the main sheet
    table = None
    if ws._tables:
        table = next(iter(ws._tables.values()))

    # Find the last row with data
    start_row = ws.max_row + 1

    # Write new records
    for row_idx, row in enumerate(df_nuevos.itertuples(index=False), start=start_row):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Update the table range if a table exists
    if table is not None:
        end_col = get_column_letter(len(df_nuevos.columns))
        end_row = ws.max_row
        table.ref = f"A1:{end_col}{end_row}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def write_cae_excel_with_pivot(
    df_nuevos: pd.DataFrame, df_actual: pd.DataFrame
) -> bytes:
    """Test utility: create a workbook with new records, data, and a native pivot.

    This is a test utility for generating sample files with native pivots.
    The main application flow uses append_new_records_to_report instead.
    """
    from pivot_builder import add_native_pivot

    cols = resolve_cae_pivot_columns(df_actual)

    key_col = df_actual.columns[0]
    drop_cols = [key_col] if key_col not in cols.values() else []
    df_data = df_actual.drop(columns=drop_cols)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_nuevos.to_excel(writer, sheet_name=REPORT_SHEET, index=False)
        _add_formatted_table(writer.sheets[REPORT_SHEET], df_nuevos, "RegistrosNuevos")
        df_data.to_excel(writer, sheet_name=DATA_SHEET, index=False)
        writer.book.create_sheet(PIVOT_SHEET)

    return add_native_pivot(
        output.getvalue(),
        df_data=df_data,
        source_sheet=DATA_SHEET,
        pivot_sheet=PIVOT_SHEET,
        row_field=cols["ETAPA SATCHMO"],
        col_field=cols["Mes"],
        value_field=cols["PRODUCTO"],
        filter_field=cols["FECHA CREA"],
    )


st.title("Reporte CAE")

col1, col2 = st.columns(2)

df_anterior = None
df_actual = None
actual_bytes = None
key_anterior_col = None

with col1:
    st.header("Reporte día anterior")
    file_anterior = st.file_uploader(
        "Subir reporte día anterior",
        type=["xlsx", "xls"],
        key="cae_anterior",
    )
    if file_anterior is not None:
        try:
            with st.spinner("Leyendo reporte día anterior..."):
                df_anterior = pd.read_excel(file_anterior, sheet_name=REPORT_SHEET)
            llave_matches = [
                c for c in df_anterior.columns if str(c).lower() == "llave"
            ]
            if llave_matches:
                candidate_col = llave_matches[0]
                candidate_values = df_anterior[candidate_col]
                valid_values = candidate_values.notna() & candidate_values.astype(
                    str
                ).str.strip().ne("")
                if valid_values.any():
                    key_anterior_col = candidate_col
            if key_anterior_col is None:
                llave_values = df_anterior["OPERACIÓN"].astype(str) + df_anterior[
                    "ETAPA SATCHMO"
                ].astype(str)
                if llave_matches:
                    key_anterior_col = llave_matches[0]
                    df_anterior[key_anterior_col] = llave_values
                else:
                    key_anterior_col = "Llave"
                    df_anterior.insert(0, key_anterior_col, llave_values)
                st.info(
                    "La columna 'Llave' del reporte día anterior estaba vacía o "
                    "no existía; se generó con OPERACIÓN + ETAPA SATCHMO."
                )
            st.success(f"Archivo cargado: {file_anterior.name}")
            st.dataframe(df_anterior.head())
        except Exception as e:
            df_anterior = None
            st.error(f"Error al leer el reporte día anterior: {e}")
            st.info(
                f"El archivo debe contener la hoja '{REPORT_SHEET}' y, si no "
                "incluye la columna 'Llave', las columnas 'OPERACIÓN' y "
                "'ETAPA SATCHMO'."
            )

with col2:
    st.header("Reporte actual")
    file_actual = st.file_uploader(
        "Subir reporte actual",
        type=["xlsx", "xls"],
        key="cae_actual",
    )
    if file_actual is not None:
        try:
            with st.spinner("Leyendo reporte actual..."):
                actual_bytes = file_actual.getvalue()
                df_actual = pd.read_excel(
                    io.BytesIO(actual_bytes), sheet_name=REPORT_SHEET
                )
            key_col = df_actual.columns[0]
            df_actual[key_col] = df_actual["OPERACIÓN"].astype(str) + df_actual[
                "ETAPA SATCHMO"
            ].astype(str)
            st.success(f"Archivo cargado: {file_actual.name}")
            st.dataframe(df_actual.head())
        except Exception as e:
            df_actual = None
            actual_bytes = None
            st.error(f"Error al leer el reporte actual: {e}")
            st.info(
                f"El archivo debe contener la hoja '{REPORT_SHEET}' y las "
                "columnas 'OPERACIÓN' y 'ETAPA SATCHMO'."
            )

if df_anterior is not None and df_actual is not None:
    st.divider()
    try:
        keys_anterior = set(df_anterior[key_anterior_col].astype(str))
        key_actual_col = df_actual.columns[0]
        df_nuevos = df_actual[
            ~df_actual[key_actual_col].astype(str).isin(keys_anterior)
        ]
        df_nuevos = df_nuevos.drop(columns=[key_actual_col])
        df_nuevos["OPERACIÓN"] = df_nuevos["OPERACIÓN"].astype(str)
        df_nuevos["CÓDIGO TRÁMITE"] = df_nuevos["CÓDIGO TRÁMITE"].astype(str)
        df_nuevos["FECHA SATCHMO"] = pd.to_datetime(
            df_nuevos["FECHA SATCHMO"], errors="coerce"
        ).dt.strftime("%d-%m-%Y")
        df_nuevos["FECHA CREA"] = pd.to_datetime(
            df_nuevos["FECHA CREA"], errors="coerce"
        ).dt.strftime("%d-%m-%Y")
        df_nuevos["RUT DEUDOR"] = pd.to_numeric(
            df_nuevos["RUT DEUDOR"], errors="coerce"
        )
        st.subheader("Registros nuevos en reporte actual")
        st.write(f"Total registros nuevos: {len(df_nuevos)}")
        st.dataframe(df_nuevos.head())

        st.subheader("Registros nuevos en reporte actual")
        st.write(f"Total registros nuevos: {len(df_nuevos)}")
        st.dataframe(df_nuevos.head())

        if df_nuevos.empty:
            st.info("No hay registros nuevos entre ambos reportes.")
        else:
            st.info(
                "El archivo descargado conserva el reporte actual completo "
                "(incluyendo su tabla dinámica en la hoja 'Hoja3') y añade "
                "los nuevos registros a la hoja principal para que la tabla "
                "dinámica los refleje al actualizar."
            )
            if actual_bytes is not None:
                try:
                    # Regenerate Llave for new records to match main sheet structure
                    key_col = df_actual.columns[0]
                    df_nuevos[key_col] = df_nuevos["OPERACIÓN"].astype(str) + df_nuevos[
                        "ETAPA SATCHMO"
                    ].astype(str)

                    # Get the column order from the main sheet in the uploaded workbook
                    # We need to read the main sheet structure from actual_bytes
                    import openpyxl

                    wb_temp = openpyxl.load_workbook(io.BytesIO(actual_bytes))
                    ws_main = wb_temp[REPORT_SHEET]
                    main_cols = [
                        cell.value
                        for cell in next(ws_main.iter_rows(min_row=1, max_row=1))
                    ]

                    # Ensure df_nuevos has the same columns in the same order
                    df_nuevos = df_nuevos.reindex(columns=main_cols)

                    excel_bytes = append_new_records_to_report(actual_bytes, df_nuevos)
                except Exception as e:
                    st.warning(
                        f"No se pudo preservar la tabla dinámica ({e}); "
                        "el archivo solo contendrá los registros nuevos."
                    )
                    excel_bytes = write_cae_nuevos_excel(df_nuevos)
            else:
                excel_bytes = write_cae_nuevos_excel(df_nuevos)

            st.download_button(
                label="Descargar registros nuevos",
                data=excel_bytes,
                file_name="registros_nuevos_cae.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )
    except Exception as e:
        st.error(f"Error al comparar los reportes: {e}")
        st.info(
            "Verifica que ambos reportes tengan las columnas 'OPERACIÓN', "
            "'CÓDIGO TRÁMITE', 'FECHA SATCHMO', 'FECHA CREA' y 'RUT DEUDOR'."
        )
elif file_anterior is not None or file_actual is not None:
    st.info("Carga ambos reportes para comparar y detectar registros nuevos.")
