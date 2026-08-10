import io
from datetime import datetime, timedelta

import pandas as pd
import streamlit as st

MESES_ESPANOL = {
    1: "ene",
    2: "feb",
    3: "mar",
    4: "abr",
    5: "may",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "sep",
    10: "oct",
    11: "nov",
    12: "dic",
}

# Feriados de Chile (mes, día) - se aplican a cualquier año
FERIADOS_CHILE = [
    (1, 1),  # Año Nuevo
    (4, 3),  # Viernes Santo
    (4, 4),  # Sábado Santo
    (5, 1),  # Día Nacional del Trabajo
    (5, 21),  # Día de las Glorias Navales
    (6, 21),  # Día Nacional de los Pueblos Indígenas
    (6, 29),  # San Pedro y San Pablo
    (7, 16),  # Día de la Virgen del Carmen
    (8, 15),  # Asunción de la Virgen
    (9, 18),  # Independencia Nacional
    (9, 19),  # Día de las Glorias del Ejército
    (10, 12),  # Encuentro de Dos Mundos
    (10, 31),  # Día de las Iglesias Evangélicas y Protestantes
    (11, 1),  # Día de Todos los Santos
    (12, 8),  # Inmaculada Concepción
    (12, 25),  # Navidad
]


def es_dia_habil(fecha: datetime.date) -> bool:
    """Verifica si una fecha es día hábil (no fin de semana ni feriado)."""
    # Verificar si es fin de semana
    if fecha.weekday() >= 5:  # 5=sábado, 6=domingo
        return False

    # Verificar si es feriado
    if (fecha.month, fecha.day) in FERIADOS_CHILE:
        return False

    return True


def obtener_proximo_dia_habil(fecha_inicio: datetime.date) -> datetime.date:
    """Obtiene el próximo día hábil desde una fecha dada."""
    proximo_dia = fecha_inicio + timedelta(days=1)

    # Avanzar hasta encontrar un día hábil
    while not es_dia_habil(proximo_dia):
        proximo_dia += timedelta(days=1)

    return proximo_dia


st.title("Compromisos")

uploaded_file = st.file_uploader(
    "Cargar archivo Excel",
    type=["xlsx", "xls"],
    key="compromisos_uploader",
)

if uploaded_file is not None:
    try:
        # Leer el archivo sin dtype=str para que las fechas se parseen automáticamente
        with st.spinner("Leyendo archivo..."):
            df = pd.read_excel(uploaded_file)

        # Eliminar columnas L, M, N, O (posiciones 11, 12, 13, 14)
        columnas_a_eliminar = [11, 12, 13, 14]
        df = df.drop(df.columns[columnas_a_eliminar], axis=1)

        # Mapeo de columnas (clave: inicio del nombre en minúsculas, valor: nombre final exacto)
        mapeo_columnas = {
            "nombre empresa": "EMPRESA",
            "cobra": "Sigla_Cobra",
            "rut clte": "  Rut_Cliente",
            "operacion": "Operación ",
            "monto a pago": "Monto  ",
            "tipo pago": "   Tipo_de_Pago ",
            "forma de pago": "Modo : presencial/boton de pago ",
            "fecha de compromiso": "FECHA DE COMPR.",
        }

        # Construir diccionario de rename por coincidencia parcial
        rename_dict = {}
        for col in df.columns:
            col_lower = col.strip().lower()
            for clave, nuevo_nombre in mapeo_columnas.items():
                if col_lower.startswith(clave):
                    rename_dict[col] = nuevo_nombre
                    break

        df_renombrado = df.rename(columns=rename_dict)

        # Filtrar solo las columnas mapeadas, con FECHA DE COMPR. al final
        columnas_finales = [
            col for col in rename_dict.values() if col != "FECHA DE COMPR."
        ]
        columnas_finales.append("FECHA DE COMPR.")
        df_final = df_renombrado[columnas_finales]

        # Calcular próximo día hábil (considera fines de semana y feriados)
        hoy = datetime.now().date()
        proximo_dia = obtener_proximo_dia_habil(hoy)

        # La columna FECHA DE COMPR. ya viene como datetime64[ns]
        fecha_col = "FECHA DE COMPR."

        # Filtrar por fecha
        df_filtrado = df_final[df_final[fecha_col].dt.date == proximo_dia].copy()

        st.success("Archivo cargado correctamente")

        st.subheader("Vista previa de los datos originales")
        st.dataframe(df.head())
        st.write(f"Total de filas: {len(df)}")
        st.write(f"Total de columnas: {len(df.columns)}")

        st.subheader(f"Compromisos para el {proximo_dia.strftime('%d/%m/%Y')}")
        if len(df_filtrado) > 0:
            st.dataframe(df_filtrado)
            st.write(
                f"Total de registros para el {proximo_dia.strftime('%d/%m/%Y')}: {len(df_filtrado)}"
            )

            # Botón de descarga
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_filtrado.to_excel(writer, index=False, sheet_name="DEMO")

                from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

                worksheet = writer.sheets["DEMO"]

                # Formato para encabezados: fondo rojo, letras blancas, negrita, Calibri 10
                header_fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
                header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

                # Bordes para todas las celdas
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # Fuente para Tipo_de_Pago: negrita y rojo
                tipo_pago_font = Font(
                    name="Aptos Narrow", size=11, bold=True, color="FF0000"
                )

                # Fuente para Modo: negrita y negro
                modo_font = Font(
                    name="Aptos Narrow", size=11, bold=True, color="000000"
                )

                # Fuente para Fecha: negrita y rojo
                fecha_font = Font(
                    name="Aptos Narrow", size=11, bold=True, color="FF0000"
                )

                # Fondo celeste para Modo
                modo_fill = PatternFill(
                    start_color="CCFFFF", end_color="CCFFFF", fill_type="solid"
                )

                # Fuente normal para datos
                data_font = Font(name="Aptos Narrow", size=11)

                # Aplicar formato a la primera fila (encabezados)
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                # Alto de fila de encabezados
                worksheet.row_dimensions[1].height = 21.60

                # Encontrar índice de columnas especiales
                col_operacion = None
                col_monto = None
                col_tipo_pago = None
                col_modo = None
                col_fecha = None

                for idx, cell in enumerate(worksheet[1], 1):
                    col_name = cell.value.strip().lower() if cell.value else ""
                    if "operación" in col_name:
                        col_operacion = idx
                    elif "monto" in col_name:
                        col_monto = idx
                    elif "tipo_de_pago" in col_name:
                        col_tipo_pago = idx
                    elif "modo" in col_name:
                        col_modo = idx
                        # Aplicar formato especial al encabezado de Modo
                        cell.fill = modo_fill
                        cell.font = modo_font
                    elif "fecha" in col_name:
                        col_fecha = idx

                # Aplicar formato a las filas de datos
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.font = data_font
                        cell.border = thin_border

                        # Columna A (EMPRESA): alinear datos a la derecha
                        if cell.column == 1:
                            cell.alignment = Alignment(
                                horizontal="right", vertical="center"
                            )

                        # Formatear columna Operación como número sin separador de miles
                        if col_operacion and cell.column == col_operacion:
                            cell.number_format = "0"

                        # Formatear columna Monto con separador de miles y alineación izquierda
                        if col_monto and cell.column == col_monto:
                            cell.number_format = "#,##0"
                            cell.alignment = Alignment(
                                horizontal="left", vertical="center"
                            )

                        # Formatear columna Tipo_de_Pago: negrita y rojo
                        if col_tipo_pago and cell.column == col_tipo_pago:
                            cell.font = tipo_pago_font

                        # Formatear fecha sin hora
                        if col_fecha and cell.column == col_fecha:
                            cell.number_format = "d-mmm"
                            cell.font = fecha_font
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )

                # Anchos de columnas fijos
                anchos_columnas = {
                    "A": 11.78,
                    "B": 11.78,
                    "C": 11.78,
                    "D": 13.00,
                    "E": 23.56,
                    "F": 13.11,
                    "G": 27.22,
                    "H": 15.00,
                }

                for col_letter, width in anchos_columnas.items():
                    worksheet.column_dimensions[col_letter].width = width

            excel_data = output.getvalue()

            dia_actual = proximo_dia.day
            mes_actual = MESES_ESPANOL[proximo_dia.month]
            anio_actual = proximo_dia.year
            nombre_archivo = f"INGRESO DE COMPROMISOS DIARIOS {proximo_dia.strftime('%d-%m-%Y')}.xlsx"

            st.download_button(
                label="Descargar archivo de compromisos",
                data=excel_data,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.warning(f"No hay compromisos para el {proximo_dia.strftime('%d/%m/%Y')}")
            st.write("Fechas disponibles en el archivo:")
            fechas_unicas = df_final[fecha_col].dropna().dt.date.unique()
            fechas_ordenadas = sorted(fechas_unicas)
            for fecha in fechas_ordenadas[:10]:  # Mostrar primeras 10 fechas
                st.write(f"- {fecha.strftime('%d/%m/%Y')}")
    except Exception as e:
        st.error(f"Error al leer el archivo: {e}")
else:
    st.info("Carga un archivo Excel para comenzar")
