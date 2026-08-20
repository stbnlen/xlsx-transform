"""Tests for the Reporte CAE page in pages/reporte_cae.py."""

import io
import os
import sys
import xml.etree.ElementTree as ET
import zipfile

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from pages.reporte_cae import (  # noqa: E402
    REPORT_SHEET,
    append_new_records_to_report,
    write_cae_nuevos_excel,
)


def create_sample_cae() -> pd.DataFrame:
    """Sample current-report data with the pivot fields."""
    return pd.DataFrame(
        {
            "Llave": ["", "", "", "", ""],
            "OPERACION": ["1", "2", "3", "4", "5"],
            "ETAPA SATCHMO": ["Etapa1", "Etapa1", "Etapa2", "Etapa2", "Etapa1"],
            "Mes": ["ene", "ene", "ene", "feb", "feb"],
            "PRODUCTO": ["P1", "P2", "P3", "P4", None],
            "FECHA CREA": pd.to_datetime(
                [
                    "2026-01-05",
                    "2026-01-10",
                    "2026-01-15",
                    "2026-02-15",
                    "2026-02-20",
                ]
            ),
        }
    )


def test_append_new_records_to_main_sheet():
    """New records are appended to the main data sheet, pivot parts preserved."""
    df_actual = create_sample_cae()
    df_actual["Llave"] = df_actual["OPERACION"] + df_actual["ETAPA SATCHMO"]

    # Create a base workbook with a native pivot (simulating the uploaded report)
    from pages.reporte_cae import write_cae_excel_with_pivot

    df_nuevos = df_actual.head(2)
    base_bytes = write_cae_excel_with_pivot(df_nuevos, df_actual)

    # Append new records
    data = append_new_records_to_report(base_bytes, df_nuevos)

    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = set(zf.namelist())
        # Original pivot parts are preserved
        assert "xl/pivotTables/pivotTable1.xml" in names
        assert "xl/pivotCache/pivotCacheDefinition1.xml" in names
        assert "xl/pivotCache/pivotCacheRecords1.xml" in names

        # Content types include pivot parts
        content_types = zf.read("[Content_Types].xml").decode("utf-8")
        assert "pivotCacheDefinition+xml" in content_types
        assert "pivotCacheRecords+xml" in content_types
        assert "pivotTable+xml" in content_types

        # Sheets: original report, DATOS, Hoja1 - NO new REGISTROS_NUEVOS sheet
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "REPORTES_GESTIONES_TODAS_ETAPAS" in wb.sheetnames
        assert "DATOS" in wb.sheetnames
        assert "Hoja1" in wb.sheetnames
        assert "REGISTROS_NUEVOS" not in wb.sheetnames

        # Main sheet has original (2) + new (2) = 4 records
        main_sheet = pd.read_excel(
            io.BytesIO(data), sheet_name="REPORTES_GESTIONES_TODAS_ETAPAS"
        )
        assert len(main_sheet) == 4
        assert "OPERACION" in main_sheet.columns
        assert "ETAPA SATCHMO" in main_sheet.columns
        assert "Mes" in main_sheet.columns
        assert "PRODUCTO" in main_sheet.columns

        # Verify new records were appended
        new_records = main_sheet.tail(2)
        assert new_records["OPERACION"].astype(str).tolist() == ["1", "2"]


def test_write_cae_nuevos_excel_single_sheet():
    """The new-records-only download keeps a single REPORT_SHEET sheet."""
    df_actual = create_sample_cae()
    df_nuevos = df_actual.head(3)

    data = write_cae_nuevos_excel(df_nuevos)

    root = ET.fromstring(zipfile.ZipFile(io.BytesIO(data)).read("xl/workbook.xml"))
    sheet_names = [s.get("name") for s in root.iter() if s.tag.endswith("}sheet")]
    assert sheet_names == [REPORT_SHEET]


def test_append_new_records_without_pivot_fallback():
    """If the uploaded file has no pivot parts, the fallback still works (appends to main sheet)."""
    df_actual = create_sample_cae()
    df_actual["Llave"] = df_actual["OPERACION"] + df_actual["ETAPA SATCHMO"]
    df_nuevos = df_actual.head(2)

    # Create a base workbook WITHOUT pivot parts (just a simple workbook)
    df_base = pd.DataFrame(
        {
            "OPERACION": ["1", "2", "3", "4"],
            "ETAPA SATCHMO": ["E1", "E1", "E2", "E2"],
            "Mes": ["ene", "feb", "ene", "feb"],
            "PRODUCTO": ["P1", "P2", "P3", "P4"],
            "FECHA CREA": pd.to_datetime(
                [
                    "2026-01-05",
                    "2026-02-10",
                    "2026-01-15",
                    "2026-02-20",
                ]
            ),
        }
    )
    base_bytes = io.BytesIO()
    with pd.ExcelWriter(base_bytes, engine="openpyxl") as writer:
        df_base.to_excel(
            writer, sheet_name="REPORTES_GESTIONES_TODAS_ETAPAS", index=False
        )

    data = append_new_records_to_report(base_bytes.getvalue(), df_nuevos)

    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = set(zf.namelist())
        assert "xl/pivotTables/pivotTable1.xml" not in names  # no pivot created

        # Only REPORT_SHEET exists, no REGISTROS_NUEVOS sheet
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert wb.sheetnames == ["REPORTES_GESTIONES_TODAS_ETAPAS"]
        main_sheet = pd.read_excel(
            io.BytesIO(data), sheet_name="REPORTES_GESTIONES_TODAS_ETAPAS"
        )
        assert len(main_sheet) == 6  # 4 original + 2 new
