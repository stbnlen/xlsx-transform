"""Build native Excel pivot tables by injecting OOXML pivot parts.

Excel pivot tables cannot be created with pandas/openpyxl alone, so this
module hand-crafts the pivotCacheDefinition, pivotCacheRecords and
pivotTable XML parts and injects them into an existing .xlsx archive.
The cache is marked ``refreshOnLoad="1"`` so Excel rebuilds it from the
worksheet source when the file is opened.
"""

import io
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime
from xml.sax.saxutils import escape

import pandas as pd
from openpyxl.utils import get_column_letter

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"

REL_PIVOT_CACHE_DEFINITION = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
    "pivotCacheDefinition"
)
REL_PIVOT_CACHE_RECORDS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
    "pivotCacheRecords"
)
REL_PIVOT_TABLE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/" "pivotTable"
)

CT_PIVOT_CACHE_DEFINITION = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml."
    "pivotCacheDefinition+xml"
)
CT_PIVOT_CACHE_RECORDS = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml."
    "pivotCacheRecords+xml"
)
CT_PIVOT_TABLE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml." "pivotTable+xml"
)

CACHE_ID = 100
_QUOTE_MAP = {'"': "&quot;"}
_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


class _CacheField:
    """A pivot cache field: name, value kind and sorted unique items."""

    def __init__(
        self,
        name: str,
        kind: str,
        items: list,
        has_blank: bool,
        index_map: dict,
    ) -> None:
        self.name = name
        self.kind = kind
        self.items = items
        self.has_blank = has_blank
        self.index_map = index_map


def _esc(value) -> str:
    """Escape a value for use inside an XML attribute or text node."""
    return _CONTROL_CHARS.sub("", escape(str(value), _QUOTE_MAP))


def _is_blank(value) -> bool:
    """Return True for None/NaN/empty-string values."""
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    return isinstance(value, str) and value.strip() == ""


def _to_datetime(value) -> datetime:
    """Coerce a date-like value to a plain datetime."""
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    return pd.to_datetime(value).to_pydatetime()


def _classify_series(series: pd.Series) -> str:
    """Classify a series as 'date', 'number' or 'string'."""
    if pd.api.types.is_datetime64_any_dtype(series):
        return "date"
    if pd.api.types.is_numeric_dtype(series):
        return "number"
    sample = [v for v in series.dropna().head(50)]
    if sample and all(isinstance(v, (datetime, date)) for v in sample):
        return "date"
    return "string"


def _value_key(value, kind: str):
    """Canonical key of a value for the shared-items index map."""
    if kind == "string":
        return str(value)
    if kind == "number":
        return float(value)
    return _to_datetime(value)


def _build_field(name: str, series: pd.Series) -> _CacheField:
    """Build a cache field with its sorted unique (non-blank) items."""
    kind = _classify_series(series)
    non_blank = [v for v in series if not _is_blank(v)]
    keys = {_value_key(v, kind) for v in non_blank}
    items = sorted(keys)
    has_blank = len(non_blank) < len(series)
    index_map = {key: idx for idx, key in enumerate(items)}
    return _CacheField(str(name), kind, items, has_blank, index_map)


def _format_number(value) -> str:
    """Format a number for XML without a trailing '.0' on integers."""
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return repr(number)


def _shared_items_xml(field: _CacheField) -> str:
    """Render the <sharedItems> element of a cache field."""
    if field.kind == "string":
        attrs = 'containsSemiMixedTypes="0" containsString="1"'
        inner = "".join(f'<s v="{_esc(v)}"/>' for v in field.items)
    elif field.kind == "number":
        attrs = 'containsSemiMixedTypes="0" containsNumber="1"'
        if field.items and all(float(v).is_integer() for v in field.items):
            attrs += ' containsInteger="1"'
        inner = "".join(f'<n v="{_format_number(v)}"/>' for v in field.items)
    else:
        attrs = 'containsSemiMixedTypes="0" containsDate="1"'
        inner = "".join(
            f'<d v="{v.strftime("%Y-%m-%dT%H:%M:%S")}"/>' for v in field.items
        )
    if field.has_blank:
        attrs += ' containsBlank="1"'
    return f'<sharedItems {attrs} count="{len(field.items)}">{inner}</sharedItems>'


def _cache_field_xml(field: _CacheField) -> str:
    """Render a <cacheField> element."""
    return (
        f'<cacheField name="{_esc(field.name)}" numFmtId="0" databaseField="1">'
        f"{_shared_items_xml(field)}</cacheField>"
    )


def _record_cell_xml(value, field: _CacheField) -> str:
    """Render one cell of a cache record (<x v=.../> or <m/>)."""
    if _is_blank(value):
        return "<m/>"
    return f'<x v="{field.index_map[_value_key(value, field.kind)]}"/>'


def _ole_now() -> float:
    """Current datetime as an OLE automation date."""
    delta = datetime.now() - datetime(1899, 12, 30)
    return round(delta.days + delta.seconds / 86400, 4)


def _cache_definition_xml(
    fields: list, n_records: int, source_sheet: str, source_ref: str
) -> str:
    """Render xl/pivotCache/pivotCacheDefinition1.xml."""
    cache_fields = "".join(_cache_field_xml(field) for field in fields)
    return (
        f'<pivotCacheDefinition xmlns="{NS_MAIN}" xmlns:r="{NS_DOC_REL}" '
        f'r:id="rId1" refreshedBy="xlsx-transform" refreshedDate="{_ole_now()}" '
        f'createdVersion="8" refreshedVersion="8" minRefreshableVersion="3" '
        f'recordCount="{n_records}" refreshOnLoad="1">'
        f'<cacheSource type="worksheet">'
        f'<worksheetSource ref="{source_ref}" sheet="{_esc(source_sheet)}"/>'
        f"</cacheSource>"
        f'<cacheFields count="{len(fields)}">{cache_fields}</cacheFields>'
        f"</pivotCacheDefinition>"
    )


def _cache_records_xml(df_data: pd.DataFrame, fields: list) -> str:
    """Render xl/pivotCache/pivotCacheRecords1.xml."""
    records = []
    for row in df_data.itertuples(index=False):
        cells = "".join(
            _record_cell_xml(value, field) for value, field in zip(row, fields)
        )
        records.append(f"<r>{cells}</r>")
    return (
        f'<pivotCacheRecords xmlns="{NS_MAIN}" xmlns:r="{NS_DOC_REL}" '
        f'count="{len(records)}">{"".join(records)}</pivotCacheRecords>'
    )


def _axis_pivot_field_xml(axis: str, field: _CacheField) -> str:
    """Render a <pivotField> placed on an axis, with one item per value."""
    items = "".join(f'<item x="{idx}"/>' for idx in range(len(field.items)))
    items += '<item t="default"/>'
    return (
        f'<pivotField axis="{axis}" showAll="0">'
        f'<items count="{len(field.items) + 1}">{items}</items></pivotField>'
    )


def _axis_items_xml(n_items: int) -> str:
    """Render <rowItems>/<colItems>: one <i> per value plus the grand total."""
    items = "".join(f'<i><x v="{idx}"/></i>' for idx in range(n_items))
    items += '<i t="grand"><x/></i>'
    return items


def _pivot_table_xml(
    fields: list,
    idx_filter: int,
    idx_row: int,
    idx_col: int,
    idx_value: int,
    pivot_name: str,
) -> str:
    """Render xl/pivotTables/pivotTable1.xml."""
    pivot_fields = []
    for idx, field in enumerate(fields):
        if idx == idx_filter:
            pivot_fields.append(_axis_pivot_field_xml("axisPage", field))
        elif idx == idx_row:
            pivot_fields.append(_axis_pivot_field_xml("axisRow", field))
        elif idx == idx_col:
            pivot_fields.append(_axis_pivot_field_xml("axisCol", field))
        elif idx == idx_value:
            pivot_fields.append('<pivotField dataField="1" showAll="0"/>')
        else:
            pivot_fields.append('<pivotField showAll="0"/>')

    n_row_items = len(fields[idx_row].items)
    n_col_items = len(fields[idx_col].items)
    last_col = get_column_letter(n_col_items + 2)
    last_row = n_row_items + 4
    location_ref = f"A3:{last_col}{last_row}"

    row_items = _axis_items_xml(n_row_items)
    col_items = _axis_items_xml(n_col_items)

    return (
        f'<pivotTableDefinition xmlns="{NS_MAIN}" xmlns:r="{NS_DOC_REL}" '
        f'name="{_esc(pivot_name)}" cacheId="{CACHE_ID}" applyNumberFormats="0" '
        f'applyBorderFormats="0" applyFontFormats="0" applyPatternFormats="0" '
        f'applyAlignmentFormats="0" applyWidthHeightFormats="1" '
        f'dataCaption="Valores" updatedVersion="8" '
        f'minRefreshableVersion="3" useAutoFormatting="0" itemPrintTitles="0" '
        f'indent="0" outline="1" outlineData="1" multipleFieldFilters="0">'
        f'<location ref="{location_ref}" firstHeaderRow="1" firstDataRow="2" '
        f'firstDataCol="1"/>'
        f'<pivotFields count="{len(fields)}">{"".join(pivot_fields)}</pivotFields>'
        f'<rowFields count="1"><field x="{idx_row}"/></rowFields>'
        f'<rowItems count="{n_row_items + 1}">{row_items}</rowItems>'
        f'<colFields count="1"><field x="{idx_col}"/></colFields>'
        f'<colItems count="{n_col_items + 1}">{col_items}</colItems>'
        f'<pageFields count="1">'
        f'<pageField fld="{idx_filter}" name="{_esc(fields[idx_filter].name)}"/>'
        f"</pageFields>"
        f'<dataFields count="1">'
        f'<dataField fld="{idx_value}" subtotal="count" '
        f'name="Recuento de {_esc(fields[idx_value].name)}" baseField="0" '
        f'baseItem="0"/></dataFields>'
        f'<pivotTableStyleInfo name="PivotStyleMedium9" showRowHeaders="1" '
        f'showColHeaders="1" showRowStripes="0" showColStripes="0" '
        f'showLastColumn="1"/>'
        f"</pivotTableDefinition>"
    )


def _rels_xml(relationships: list) -> str:
    """Render a .rels part from (rel_id, rel_type, target) tuples."""
    inner = "".join(
        f'<Relationship Id="{rel_id}" Type="{rel_type}" Target="{target}"/>'
        for rel_id, rel_type, target in relationships
    )
    return f'<Relationships xmlns="{NS_PKG_REL}">{inner}</Relationships>'


def _find_sheet_target(workbook_xml: bytes, rels_xml: bytes, sheet_name: str) -> str:
    """Find the worksheet target file (e.g. 'worksheets/sheet3.xml')."""
    root = ET.fromstring(workbook_xml)
    sheet_rid = None
    for sheet in root.iter(f"{{{NS_MAIN}}}sheet"):
        if sheet.get("name") == sheet_name:
            sheet_rid = sheet.get(f"{{{NS_DOC_REL}}}id")
            break
    if sheet_rid is None:
        raise ValueError(f"Hoja '{sheet_name}' no encontrada en el libro")
    rels_root = ET.fromstring(rels_xml)
    for rel in rels_root:
        if rel.get("Id") == sheet_rid:
            return rel.get("Target")
    raise ValueError(f"Relación '{sheet_rid}' no encontrada en el libro")


def add_native_pivot(
    xlsx_bytes: bytes,
    df_data: pd.DataFrame,
    source_sheet: str,
    pivot_sheet: str,
    row_field: str,
    col_field: str,
    value_field: str,
    filter_field: str,
    pivot_name: str = "TablaDinamicaCAE",
) -> bytes:
    """Inject a native Excel pivot table into an existing workbook.

    The pivot reads from ``source_sheet`` (which must already contain
    ``df_data`` starting at A1) and is placed on ``pivot_sheet``.
    """
    fields = [_build_field(column, df_data[column]) for column in df_data.columns]
    column_names = list(df_data.columns)
    idx_row = column_names.index(row_field)
    idx_col = column_names.index(col_field)
    idx_value = column_names.index(value_field)
    idx_filter = column_names.index(filter_field)

    end_col = get_column_letter(len(df_data.columns))
    source_ref = f"A1:{end_col}{len(df_data) + 1}"

    parts = {
        "xl/pivotCache/pivotCacheDefinition1.xml": _cache_definition_xml(
            fields, len(df_data), source_sheet, source_ref
        ),
        "xl/pivotCache/pivotCacheRecords1.xml": _cache_records_xml(df_data, fields),
        "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels": _rels_xml(
            [("rId1", REL_PIVOT_CACHE_RECORDS, "pivotCacheRecords1.xml")]
        ),
        "xl/pivotTables/pivotTable1.xml": _pivot_table_xml(
            fields, idx_filter, idx_row, idx_col, idx_value, pivot_name
        ),
        "xl/pivotTables/_rels/pivotTable1.xml.rels": _rels_xml(
            [
                (
                    "rId1",
                    REL_PIVOT_CACHE_DEFINITION,
                    "../pivotCache/pivotCacheDefinition1.xml",
                )
            ]
        ),
    }

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zip_in:
        entries = {name: zip_in.read(name) for name in zip_in.namelist()}

    workbook_xml = entries["xl/workbook.xml"]
    rels_xml = entries["xl/_rels/workbook.xml.rels"]
    sheet_target = _find_sheet_target(workbook_xml, rels_xml, pivot_sheet)

    content_types = entries["[Content_Types].xml"].decode("utf-8")
    overrides = (
        f'<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml" '
        f'ContentType="{CT_PIVOT_CACHE_DEFINITION}"/>'
        f'<Override PartName="/xl/pivotCache/pivotCacheRecords1.xml" '
        f'ContentType="{CT_PIVOT_CACHE_RECORDS}"/>'
        f'<Override PartName="/xl/pivotTables/pivotTable1.xml" '
        f'ContentType="{CT_PIVOT_TABLE}"/>'
    )
    entries["[Content_Types].xml"] = content_types.replace(
        "</Types>", overrides + "</Types>"
    ).encode("utf-8")

    rels_text = rels_xml.decode("utf-8")
    existing_ids = [int(match) for match in re.findall(r'Id="rId(\d+)"', rels_text)]
    cache_rel_id = f"rId{max(existing_ids) + 1}"
    entries["xl/_rels/workbook.xml.rels"] = rels_text.replace(
        "</Relationships>",
        f'<Relationship Id="{cache_rel_id}" Type="{REL_PIVOT_CACHE_DEFINITION}" '
        f'Target="pivotCache/pivotCacheDefinition1.xml"/></Relationships>',
    ).encode("utf-8")

    workbook_text = workbook_xml.decode("utf-8")
    entries["xl/workbook.xml"] = workbook_text.replace(
        "</workbook>",
        f'<pivotCaches><pivotCache cacheId="{CACHE_ID}" r:id="{cache_rel_id}"/>'
        f"</pivotCaches></workbook>",
    ).encode("utf-8")

    sheet_rels_name = f"xl/worksheets/_rels/{sheet_target.split('/')[-1]}.rels"
    sheet_rel = [("rId1", REL_PIVOT_TABLE, "../pivotTables/pivotTable1.xml")]
    if sheet_rels_name in entries:
        sheet_rels_text = entries[sheet_rels_name].decode("utf-8")
        entries[sheet_rels_name] = sheet_rels_text.replace(
            "</Relationships>",
            f'<Relationship Id="{sheet_rel[0][0]}" Type="{sheet_rel[0][1]}" '
            f'Target="{sheet_rel[0][2]}"/></Relationships>',
        ).encode("utf-8")
    else:
        entries[sheet_rels_name] = _rels_xml(sheet_rel).encode("utf-8")

    for part_name, part_xml in parts.items():
        entries[part_name] = part_xml.encode("utf-8")

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zip_out:
        for name, data in entries.items():
            zip_out.writestr(name, data)
    return output.getvalue()
